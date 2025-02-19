from flask import Flask, request, jsonify, render_template, send_file
import pandas as pd
import os
import chardet
from openpyxl import Workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)

UPLOAD_FOLDER = "uploads"
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def detect_encoding(file):
    """Detect encoding of a file to avoid UTF-8 errors."""
    raw_data = file.read(50000)
    file.seek(0)
    result = chardet.detect(raw_data)
    return result['encoding']

def load_file(file):
    """Load file based on its format (CSV or Excel)."""
    filename = file.filename.lower()

    if filename.endswith('.csv'):
        encoding = detect_encoding(file)
        return pd.read_csv(file, encoding=encoding)
    elif filename.endswith(('.xls', '.xlsx')):
        return pd.read_excel(file, engine="openpyxl")
    else:
        raise ValueError("Unsupported file format. Please upload a CSV or Excel file. Copy provided format.")

@app.route("/")
def home():
    return render_template("index.html")

@app.route("/check_compliance", methods=["POST"])
def check_compliance():
    try:
        if 'stock_file' not in request.files or 'outbound_file' not in request.files:
            return jsonify({"error": "Both stock_file and outbound_file are required."}), 400

        stock_file = request.files['stock_file']
        outbound_file = request.files['outbound_file']

        df_stock = load_file(stock_file)
        df_outbound = load_file(outbound_file)

        df_stock.columns = df_stock.columns.str.strip()
        df_outbound.columns = df_outbound.columns.str.strip()

        required_columns = {'SKU', 'Batch No', 'Expiry Date', 'Quantity', 'Storage Location'}
        if not required_columns.issubset(df_stock.columns) or not required_columns.issubset(df_outbound.columns):
            return jsonify({"error": "Required columns are missing in one of the files"}), 400

        df_stock['Expiry Date'] = pd.to_datetime(df_stock['Expiry Date'], errors='coerce')
        df_outbound['Expiry Date'] = pd.to_datetime(df_outbound['Expiry Date'], errors='coerce')

        df_stock.sort_values(by=['SKU', 'Expiry Date'], inplace=True)

        df_outbound['Compliance Status'] = 'Compliant'
        df_outbound['Compliance Indicator'] = '🟢'

        for index, outbound_row in df_outbound.iterrows():
            sku = outbound_row['SKU']
            outbound_expiry = outbound_row['Expiry Date']
            stock_entries = df_stock[df_stock['SKU'] == sku]

            if not stock_entries.empty:
                oldest_expiry = stock_entries.iloc[0]['Expiry Date']
                if outbound_expiry > oldest_expiry:
                    df_outbound.at[index, 'Compliance Status'] = 'Non-Compliant'
                    df_outbound.at[index, 'Compliance Indicator'] = '🔴'

        df_outbound['Expiry Date'] = df_outbound['Expiry Date'].dt.strftime('%d-%m-%Y')

        compliance_data = df_outbound[['SKU', 'Batch No', 'Expiry Date', 'Quantity', 'Storage Location', 'Compliance Status', 'Compliance Indicator']].to_dict(orient='records')

        compliant_count = (df_outbound['Compliance Status'] == 'Compliant').sum()
        non_compliant_count = (df_outbound['Compliance Status'] == 'Non-Compliant').sum()
        overall_status = "Compliant" if non_compliant_count == 0 else "Non-Compliant"

        # Generate Excel Report with Colors
        report_filename = os.path.join(UPLOAD_FOLDER, "compliance_report.xlsx")
        save_report_with_colors(df_outbound, report_filename)

        return jsonify({
            "message": f"Compliance Check Completed. Out of {len(df_outbound)} SKU Batches, {compliant_count} are compliant, and {non_compliant_count} are non-compliant.",
            "compliance_report": compliance_data,
            "overall_compliance_status": overall_status,
            "download_report_url": "/download_report"
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

def save_report_with_colors(df, filename):
    """Save DataFrame to Excel with colored Compliance Status."""
    wb = Workbook()
    ws = wb.active
    ws.append(df.columns.tolist())

    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for _, row in df.iterrows():
        row_data = row.tolist()
        ws.append(row_data)

        last_row = ws.max_row
        compliance_status = row['Compliance Status']
        if compliance_status == "Compliant":
            ws[f"F{last_row}"].fill = green_fill  # Column F: Compliance Status
        else:
            ws[f"F{last_row}"].fill = red_fill  # Column F: Compliance Status

    wb.save(filename)

@app.route("/download_report", methods=["GET"])
def download_report():
    report_filename = os.path.join(UPLOAD_FOLDER, "compliance_report.xlsx")
    if os.path.exists(report_filename):
        return send_file(report_filename, as_attachment=True)
    else:
        return jsonify({"error": "Compliance report not found."}), 404

if __name__ == "__main__":
    app.run(debug=True)
